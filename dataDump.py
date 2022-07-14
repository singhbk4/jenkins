def dataread():
 from openpyxl import load_workbook
 wb = load_workbook('variable.xlsx')
 ws = wb.active
 row = 2
 epgdict = {}
 while ws.cell(row=row, column = 1).value !=None:
    tmp1 = {}
    tmp2 = {}
    vlandicttmp = {}
    epg = ws.cell(row=row, column = 1).value

    col = 2
    while ws.cell(row=1, column = col).value !=None:
        tmp2 = {ws.cell(row=1, column = col).value:ws.cell(row=row, column = col).value}
        tmp1.update(tmp2)
        col =  col+1

    vlandicttmp = {epg:tmp1}
    epgdict.update(vlandicttmp)
    row=row+1

 return epgdict






